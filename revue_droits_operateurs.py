#!/usr/bin/env python3
"""
Revue des droits opérateurs SACRE / PKI / KSTAMP
Automatisation du contrôle SSI hebdomadaire (réf. 20AA-MM-JJ-Revue-Droits-Opérateurs)

Usage :
  python revue_droits_operateurs.py
    --lir   "Lien entre Individu et Rôle.xlsx"
    --sacre SACRE_audit_des_droits_operateurs.csv
    --pki   file_output_PKI.txt
    --kstamp file_output_KSTAMP.txt
    [--output 2026-05-13-Revue-Droits-Operateurs.xlsx]
    [--date  2026-05-13]
"""

import argparse
import csv
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── Rôles de confiance attendus ───────────────────────────────────────────────

ROLE_ADMIN     = "Ingénieur / Administrateur système"
ROLE_OPERATEUR = "Opérateur"

# Fonctions SACRE déclenchant un contrôle de rôle
SACRE_SENSITIVE = {
    "demande de cle de test",
    "gestion des droits operateur",
    "gestion des instances",
    "gestion des roles utilisateur",
}

# Comptes KSTAMP non-nominatifs à ne pas contrôler
KSTAMP_SKIP = {"admin kstamp operateur", "admin kstamp backup operateur"}


# ── Normalisation des noms ────────────────────────────────────────────────────

def norm(s: str) -> str:
    """Minuscules + suppression accents + strip."""
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s).strip())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


def strip_operateur(name: str) -> str:
    """'Hamza RAHMOUNI OPERATEUR' → 'Hamza RAHMOUNI'"""
    return re.sub(r"\s+OPERATEUR\s*$", "", name.strip(), flags=re.IGNORECASE).strip()


# ── Chargement du LIR ─────────────────────────────────────────────────────────

def load_lir(path: str) -> tuple[dict, dict]:
    """
    Charge le fichier LIR (Excel, feuille 'LIR', données à partir de la ligne 7).
    Retourne :
      - lir_direct  : norm(nom_prenom) → [(rôle, domaine), ...]
      - lir_wordset : frozenset(mots) → [norm_key, ...]  (matching insensible à l'ordre)
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["LIR"]

    lir_direct  = defaultdict(list)
    lir_wordset = defaultdict(list)

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 6:          # lignes 0-5 : titre + entête
            continue
        if not row[0]:     # ligne vide (statut absent)
            continue
        nom_prenom, role, domaine = row[1], row[6], row[7]
        if not nom_prenom or not role or not domaine:
            continue

        k = norm(str(nom_prenom))
        entry = (str(role).strip(), str(domaine).strip())
        lir_direct[k].append(entry)

        ws_key = frozenset(k.split())
        if k not in lir_wordset[ws_key]:
            lir_wordset[ws_key].append(k)

    wb.close()
    return lir_direct, lir_wordset


def lir_lookup(lir_direct: dict, lir_wordset: dict, name: str) -> list:
    """
    Recherche une personne dans le LIR.
    Essaie dans l'ordre :
      1. correspondance directe (normalisée)
      2. inversion des 2 mots  (« NOM Prénom » vs « Prénom NOM »)
      3. ensemble de mots       (insensible à l'ordre, 1 seul candidat)
    Retourne la liste [(rôle, domaine)] ou [].
    """
    k = norm(name)
    if k in lir_direct:
        return lir_direct[k]

    words = k.split()
    if len(words) == 2:
        k_rev = f"{words[1]} {words[0]}"
        if k_rev in lir_direct:
            return lir_direct[k_rev]

    ws_key = frozenset(words)
    candidates = lir_wordset.get(ws_key, [])
    if len(candidates) == 1:
        return lir_direct[candidates[0]]

    return []


def is_conformant(entries: list, allowed_roles: list, allowed_domains: list) -> bool:
    """True si au moins une entrée LIR a un rôle ET un domaine autorisés."""
    for role, domain in entries:
        role_ok   = any(norm(role)   == norm(ar) for ar in allowed_roles)
        domain_ok = any(norm(domain) == norm(ad) for ad in allowed_domains)
        if role_ok and domain_ok:
            return True
    return False


def build_status(entries: list, allowed_roles: list, allowed_domains: list) -> str:
    if not entries:
        return "NON TROUVÉ LIR"
    return "CONFORME" if is_conformant(entries, allowed_roles, allowed_domains) else "RÔLE INADÉQUAT"


# ── Parsing SACRE ─────────────────────────────────────────────────────────────

def parse_sacre(path: str) -> list:
    """
    CSV « ; » — colonnes : SACREUSERNUMBER;NAME;FIRSTNAME;...;LABEL_FR
    Retourne la liste des personnes ayant ≥1 fonction sensible,
    avec leur nom complet (FIRSTNAME + NAME) et l'ensemble de leurs fonctions sensibles.
    """
    persons = {}   # sacreusernumber → dict

    # Tente UTF-8 d'abord ; si le fichier est en cp1252 (Windows), bascule
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            open(path, encoding=enc).read()
            break
        except UnicodeDecodeError:
            continue
    else:
        enc = "utf-8-sig"

    with open(path, encoding=enc, errors="replace") as f:
        reader = csv.reader(f, delimiter=";")
        header = None
        for row in reader:
            if len(row) < 5:
                continue
            row = [c.strip().strip('"') for c in row]
            if header is None:
                if row[0].upper() == "SACREUSERNUMBER":
                    header = row
                continue

            sacrenum   = row[0]
            lastname   = row[1].strip()
            firstname  = row[2].strip()
            label_raw  = row[9].strip() if len(row) > 9 else ""

            full_name = f"{firstname} {lastname}".strip()

            if sacrenum not in persons:
                persons[sacrenum] = {
                    "sacreusernumber": sacrenum,
                    "full_name": full_name,
                    "all_functions": set(),
                }
            persons[sacrenum]["all_functions"].add(label_raw)

    result = []
    for p in persons.values():
        sensitive = {f for f in p["all_functions"] if norm(f) in SACRE_SENSITIVE}
        if sensitive:
            result.append({**p, "sensitive_functions": sensitive})

    return result


# ── Parsing PKI ───────────────────────────────────────────────────────────────

def parse_pki(path: str) -> dict:
    """
    Fichier texte pipe-séparé (sortie psql).
    Section « Membres des Groupes » → group | DN du membre.
    Retourne dict : norm(prénom nom) → {name, groups: set}
    Les doublons (multi-certificats) sont agrégés par personne.
    """
    with open(path, encoding="utf-8", errors="replace") as f:
        content = f.read()

    m = re.search(
        r"Membres des Groupes\s*=+\s*(.*?)(?:\(\d+ rows\)|\Z)",
        content, re.DOTALL
    )
    if not m:
        raise ValueError("Section 'Membres des Groupes' introuvable dans le fichier PKI")

    persons = {}
    for line in m.group(1).splitlines():
        line = line.strip()
        if not line:
            continue

        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 2:
            continue

        # Ignorer entête et séparateurs
        if re.match(r"^[-+\s]+$", parts[0]) or parts[0].lower() in ("groupe", ""):
            continue

        group = parts[0]
        dn    = parts[1]
        if not group or not dn:
            continue

        cn_m = re.search(r"CN=([^,]+)", dn)
        if not cn_m:
            continue

        name = strip_operateur(cn_m.group(1))
        k = norm(name)
        if k not in persons:
            persons[k] = {"name": name, "groups": set()}
        persons[k]["groups"].add(group)

    return persons


# ── Parsing KSTAMP ────────────────────────────────────────────────────────────

_KSTAMP_FLAGS = [
    "workspace_management", "ds_management", "key_management",
    "user_management", "audit_management", "admin_management",
]


def parse_kstamp(path: str) -> list:
    """
    Fichier texte pipe-séparé (sortie psql avec entête shell).
    Colonnes : admin_dn | name | ws | ds | key | user | audit | admin  (valeurs t/f)
    Retourne la liste des utilisateurs nominatifs avec leurs flags et profil.
    """
    with open(path, encoding="utf-8", errors="replace") as f:
        lines = f.readlines()

    records = {}   # norm(name) → dict

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        parts = [p.strip() for p in stripped.split("|")]
        p0 = parts[0]

        # Ignorer : ligne de commande shell, entête, séparateur, compteur
        if p0.startswith("+"):
            continue
        if re.match(r"^[-\s]+$", p0):
            continue
        if p0.lower() == "admin_dn":
            continue
        if re.match(r"^\(\d+ rows\)", stripped):
            continue
        if len(parts) < 7:
            continue

        name_raw = parts[1]
        if not name_raw or norm(name_raw) in KSTAMP_SKIP:
            continue

        name = strip_operateur(name_raw)
        k = norm(name)

        flags = {
            flag: (len(parts) > (2 + i) and parts[2 + i].lower() == "t")
            for i, flag in enumerate(_KSTAMP_FLAGS)
        }

        # Profil : admin si l'un des 4 droits opérationnels est activé
        if any(flags[f] for f in ["workspace_management", "ds_management",
                                   "key_management", "user_management"]):
            profile = "admin"
        elif flags["audit_management"]:
            profile = "monitoring"
        else:
            profile = "admin_only"    # uniquement admin_management → pas de contrôle

        records[k] = {"name": name, "profile": profile, **flags}

    return list(records.values())


# ── Moteur d'analyse ──────────────────────────────────────────────────────────

def _disc(systeme, name, detail, status, entries):
    return {"systeme": systeme, "name": name, "detail": detail,
            "status": status, "roles_lir": entries}


def analyze_sacre(data, lir_direct, lir_wordset):
    results, disc = [], []
    for p in data:
        name    = p["full_name"]
        entries = lir_lookup(lir_direct, lir_wordset, name)
        status  = build_status(entries, [ROLE_ADMIN, ROLE_OPERATEUR], ["OSC"])
        results.append({"name": name, "functions": p["sensitive_functions"],
                        "status": status, "roles_lir": entries})
        if status != "CONFORME":
            disc.append(_disc("SACRE", name,
                              ", ".join(sorted(p["sensitive_functions"])),
                              status, entries))
    return results, disc


def analyze_pki(pki_data, lir_direct, lir_wordset):
    results, disc = [], []
    for p in pki_data.values():
        name = p["name"]
        for group in sorted(p["groups"]):
            if group == "SERVER":
                results.append({"name": name, "group": group,
                                 "status": "N/A (compte serveur)", "roles_lir": []})
                continue
            # RUN : admin OU opérateur ; tous les autres : admin uniquement
            allowed = [ROLE_ADMIN, ROLE_OPERATEUR] if group == "RUN" else [ROLE_ADMIN]
            entries = lir_lookup(lir_direct, lir_wordset, name)
            status  = build_status(entries, allowed, ["OSC"])
            results.append({"name": name, "group": group,
                             "status": status, "roles_lir": entries})
            if status != "CONFORME":
                disc.append(_disc("PKI", name,
                                  f"Groupe {group}",
                                  status, entries))
    return results, disc


def analyze_kstamp(data, lir_direct, lir_wordset):
    results, disc = [], []
    for p in data:
        name    = p["name"]
        profile = p["profile"]
        flags   = {f: p[f] for f in _KSTAMP_FLAGS}

        if profile == "admin_only":
            results.append({"name": name, "profile": profile,
                             "status": "N/A (admin_management seul)", "roles_lir": [], **flags})
            continue

        # Profil admin : ingénieur/admin seulement
        # Profil monitoring (audit seul) : ingénieur/admin OU opérateur
        allowed = [ROLE_ADMIN] if profile == "admin" else [ROLE_ADMIN, ROLE_OPERATEUR]
        entries = lir_lookup(lir_direct, lir_wordset, name)
        status  = build_status(entries, allowed, ["OSH"])
        results.append({"name": name, "profile": profile,
                         "status": status, "roles_lir": entries, **flags})
        if status != "CONFORME":
            disc.append(_disc("KSTAMP", name,
                              f"Profil {profile}",
                              status, entries))
    return results, disc


# ── Génération du rapport Excel ───────────────────────────────────────────────

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
GREY  = PatternFill("solid", fgColor="D9D9D9")
HDR_F = PatternFill("solid", fgColor="4472C4")
HDR_T = Font(color="FFFFFF", bold=True)


def _status_fill(status: str) -> PatternFill:
    if status == "CONFORME":
        return GREEN
    if status.startswith("N/A"):
        return GREY
    return RED


def _write_header(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]:
        c.fill = HDR_F
        c.font = HDR_T
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws["A2"]


def _auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 70)


def _roles_str(entries):
    return " | ".join(f"{r} / {d}" for r, d in entries) if entries else "—"


def generate_report(output_path: str, control_date: str,
                    sacre_res, sacre_disc,
                    pki_res, pki_disc,
                    kstamp_res, kstamp_disc):

    wb = openpyxl.Workbook()

    # ── Feuille SACRE ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "SACRE"
    _write_header(ws, ["Nom complet", "Fonctions sensibles détectées", "Statut", "Rôles LIR (rôle / domaine)"])
    for r in sorted(sacre_res, key=lambda x: x["name"]):
        ws.append([r["name"], ", ".join(sorted(r["functions"])),
                   r["status"], _roles_str(r["roles_lir"])])
        for c in ws[ws.max_row]:
            c.fill = _status_fill(r["status"])
    _auto_width(ws)

    # ── Feuille PKI ────────────────────────────────────────────────────────────
    ws_pki = wb.create_sheet("PKI")
    _write_header(ws_pki, ["Nom complet", "Groupe PKI", "Statut", "Rôles LIR (rôle / domaine)"])
    for r in sorted(pki_res, key=lambda x: (x["group"], x["name"])):
        ws_pki.append([r["name"], r["group"], r["status"], _roles_str(r["roles_lir"])])
        for c in ws_pki[ws_pki.max_row]:
            c.fill = _status_fill(r["status"])
    _auto_width(ws_pki)

    # ── Feuille KSTAMP ─────────────────────────────────────────────────────────
    ws_k = wb.create_sheet("KSTAMP")
    _write_header(ws_k, ["Nom", "Profil",
                          "WS", "DS", "Key", "User", "Audit", "Admin",
                          "Statut", "Rôles LIR (rôle / domaine)"])
    for r in sorted(kstamp_res, key=lambda x: x["name"]):
        f = lambda k: "oui" if r.get(k) else "non"
        ws_k.append([r["name"], r["profile"],
                     f("workspace_management"), f("ds_management"),
                     f("key_management"),       f("user_management"),
                     f("audit_management"),     f("admin_management"),
                     r["status"], _roles_str(r.get("roles_lir", []))])
        for c in ws_k[ws_k.max_row]:
            c.fill = _status_fill(r["status"])
    _auto_width(ws_k)

    # ── Feuille ÉCARTS (tickets EasyVista) ─────────────────────────────────────
    ws_e = wb.create_sheet("ÉCARTS")
    _write_header(ws_e, [
        "Système", "Nom", "Détail écart", "Statut",
        "Rôles LIR trouvés", "Objet ticket EasyVista",
        "Nomenclature", "Domaine", "Référentiel", "Sévérité", "Priorité",
    ])
    y, m = control_date[:4], control_date[5:7]
    all_disc = sacre_disc + pki_disc + kstamp_disc
    for d in all_disc:
        ticket_obj = f"[CONTROLE] {y}.{m}-Revue Droits Opérateurs - {d['name']}"
        ws_e.append([
            d["systeme"], d["name"], d["detail"], d["status"],
            _roles_str(d["roles_lir"]),
            ticket_obj,
            "SSI/Incident de sécurité",
            "APPLICATIONS SITE CENTRAL",
            "SMSI",
            "Critique",
            "Haute",
        ])
        for c in ws_e[ws_e.max_row]:
            c.fill = RED
    _auto_width(ws_e)

    # ── Feuille RÉSUMÉ (en tête de classeur) ───────────────────────────────────
    ws_s = wb.create_sheet("RÉSUMÉ", 0)
    ws_s["A1"] = f"Revue des droits opérateurs SACRE / PKI / KSTAMP — {control_date}"
    ws_s["A1"].font = Font(bold=True, size=13)
    ws_s.append([])
    _write_header(ws_s, ["Système", "Opérateurs contrôlés", "Conformes", "Écarts"])

    def _stats(results):
        ctrl = [r for r in results if not r["status"].startswith("N/A")]
        ok   = sum(1 for r in ctrl if r["status"] == "CONFORME")
        return len(ctrl), ok, len(ctrl) - ok

    for label, res in [("SACRE", sacre_res), ("PKI", pki_res),
                        ("KSTAMP", kstamp_res)]:
        t, ok, ko = _stats(res)
        ws_s.append([label, t, ok, ko])
        ws_s[ws_s.max_row][3].fill = RED if ko > 0 else GREEN

    ws_s.append([])
    ws_s.append(["Total écarts", "", "", len(all_disc)])
    ws_s[ws_s.max_row][3].font = Font(bold=True)
    ws_s[ws_s.max_row][3].fill = RED if all_disc else GREEN
    _auto_width(ws_s)

    wb.save(output_path)


# ── Point d'entrée ────────────────────────────────────────────────────────────

def main():
    # Assure que la console Windows accepte les caractères accentués
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    ap = argparse.ArgumentParser(
        description="Revue des droits opérateurs SACRE / PKI / KSTAMP"
    )
    ap.add_argument("--lir",    required=True, metavar="LIR.xlsx")
    ap.add_argument("--sacre",  required=True, metavar="SACRE.csv")
    ap.add_argument("--pki",    required=True, metavar="PKI.txt")
    ap.add_argument("--kstamp", required=True, metavar="KSTAMP.txt")
    ap.add_argument("--output", default=None)
    ap.add_argument("--date",   default=date.today().strftime("%Y-%m-%d"))
    args = ap.parse_args()

    output = args.output or f"{args.date}-Revue-Droits-Operateurs.xlsx"

    print(f"{'='*60}")
    print(f"  Revue droits opérateurs — {args.date}")
    print(f"{'='*60}")

    print("\n[1/6] Chargement LIR...")
    lir_direct, lir_wordset = load_lir(args.lir)
    print(f"      {len(lir_direct)} personnes indexées")

    print("\n[2/4] Analyse SACRE...")
    sacre_data = parse_sacre(args.sacre)
    sacre_res, sacre_disc = analyze_sacre(sacre_data, lir_direct, lir_wordset)
    print(f"      {len(sacre_res)} opérateur(s) avec fonctions sensibles, "
          f"{len(sacre_disc)} écart(s)")

    print("\n[3/4] Analyse PKI...")
    pki_data = parse_pki(args.pki)
    pki_res, pki_disc = analyze_pki(pki_data, lir_direct, lir_wordset)
    nominatifs_pki = sum(1 for r in pki_res if r["status"] != "N/A (compte serveur)")
    print(f"      {nominatifs_pki} opérateur(s) nominatifs, {len(pki_disc)} écart(s)")

    print("\n[4/4] Analyse KSTAMP...")
    kstamp_res, kstamp_disc = analyze_kstamp(
        parse_kstamp(args.kstamp), lir_direct, lir_wordset
    )
    print(f"      {len(kstamp_res)} entrée(s), {len(kstamp_disc)} écart(s)")

    # ── Synthèse des écarts ────────────────────────────────────────────────────
    all_disc = sacre_disc + pki_disc + kstamp_disc
    print(f"\n{'='*60}")
    print(f"  TOTAL ÉCARTS : {len(all_disc)}")
    print(f"{'='*60}")
    for d in all_disc:
        print(f"  [!] [{d['systeme']}] {d['name']}")
        print(f"       Statut  : {d['status']}")
        print(f"       Detail  : {d['detail']}")
        if d["roles_lir"]:
            print(f"       LIR     : {_roles_str(d['roles_lir'])}")
        print()

    # ── Rapport Excel ──────────────────────────────────────────────────────────
    print(f"Génération du rapport : {output}")
    generate_report(
        output, args.date,
        sacre_res,   sacre_disc,
        pki_res,     pki_disc,
        kstamp_res,  kstamp_disc,
    )
    print(f"Rapport généré : {output}")
    print(f"\nFeuilles : RÉSUMÉ | SACRE | PKI | KSTAMP | ÉCARTS")


if __name__ == "__main__":
    main()
