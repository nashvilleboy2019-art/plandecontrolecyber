"""
DSO-LOG-03 — Moteur d'analyse Revue des Droits Opérateurs SACRE / PKI / KSTAMP
Adapté pour utilisation dans l'app web plandecontrole.
Les fonctions de parsing acceptent du texte (str) au lieu de chemins de fichiers.
La LIR est lue depuis la base BaseLIR (SQLite) plutôt qu'un fichier Excel.
"""

import csv
import io
import re
import sqlite3
import unicodedata
from collections import defaultdict
from datetime import date

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── Constantes ────────────────────────────────────────────────────────────────

ROLE_ADMIN     = "Ingénieur / Administrateur système"
ROLE_OPERATEUR = "Opérateur"

SACRE_SENSITIVE = {
    "demande de cle de test",
    "gestion des droits operateur",
    "gestion des instances",
    "gestion des roles utilisateur",
}

KSTAMP_SKIP = {"admin kstamp operateur", "admin kstamp backup operateur"}

_KSTAMP_FLAGS = [
    "workspace_management", "ds_management", "key_management",
    "user_management", "audit_management", "admin_management",
]


# ── Normalisation ─────────────────────────────────────────────────────────────

def norm(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFD", str(s).strip())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().strip()


def strip_operateur(name: str) -> str:
    return re.sub(r"\s+OPERATEUR\s*$", "", name.strip(), flags=re.IGNORECASE).strip()


# ── Chargement LIR depuis BaseLIR SQLite ──────────────────────────────────────

def load_lir_from_db(db_path: str) -> tuple:
    """
    Charge les habilitations depuis la base BaseLIR.
    Retourne (lir_direct, lir_wordset) dans le même format que load_lir().
    """
    lir_direct  = defaultdict(list)
    lir_wordset = defaultdict(list)

    try:
        conn = sqlite3.connect(db_path)
        rows = conn.execute("""
            SELECT h.nom_prenom, r.label, d.label
            FROM habilitations h
            JOIN ref_roles r ON h.role_id = r.id
            JOIN ref_domaines d ON h.domaine_id = d.id
            WHERE h.nom_prenom IS NOT NULL
              AND r.label IS NOT NULL
              AND d.label IS NOT NULL
        """).fetchall()
        conn.close()
    except Exception:
        return lir_direct, lir_wordset

    for nom_prenom, role, domaine in rows:
        k     = norm(str(nom_prenom))
        entry = (str(role).strip(), str(domaine).strip())
        lir_direct[k].append(entry)
        ws_key = frozenset(k.split())
        if k not in lir_wordset[ws_key]:
            lir_wordset[ws_key].append(k)

    return lir_direct, lir_wordset


def lir_lookup(lir_direct: dict, lir_wordset: dict, name: str) -> list:
    k = norm(name)
    if k in lir_direct:
        return lir_direct[k]
    words = k.split()
    if len(words) == 2:
        k_rev = f"{words[1]} {words[0]}"
        if k_rev in lir_direct:
            return lir_direct[k_rev]
    ws_key     = frozenset(words)
    candidates = lir_wordset.get(ws_key, [])
    if len(candidates) == 1:
        return lir_direct[candidates[0]]
    return []


def is_conformant(entries: list, allowed_roles: list, allowed_domains: list) -> bool:
    for role, domain in entries:
        if any(norm(role) == norm(ar) for ar in allowed_roles):
            if any(norm(domain) == norm(ad) for ad in allowed_domains):
                return True
    return False


def build_status(entries: list, allowed_roles: list, allowed_domains: list) -> str:
    if not entries:
        return "NON TROUVÉ LIR"
    return "CONFORME" if is_conformant(entries, allowed_roles, allowed_domains) else "RÔLE INADÉQUAT"


# ── Parsing SACRE (texte CSV) ─────────────────────────────────────────────────

def parse_sacre(content: str) -> list:
    """
    Accepte le contenu texte d'un CSV ';'.
    Colonnes : SACREUSERNUMBER;NAME;FIRSTNAME;...;LABEL_FR
    Retourne les personnes ayant >= 1 fonction sensible.
    """
    persons = {}
    reader  = csv.reader(io.StringIO(content), delimiter=";")
    header  = None
    for row in reader:
        if len(row) < 5:
            continue
        row = [c.strip().strip('"') for c in row]
        if header is None:
            if row[0].upper() == "SACREUSERNUMBER":
                header = row
            continue
        sacrenum  = row[0]
        lastname  = row[1].strip()
        firstname = row[2].strip()
        label_raw = row[9].strip() if len(row) > 9 else ""
        full_name = f"{firstname} {lastname}".strip()
        if sacrenum not in persons:
            persons[sacrenum] = {"sacreusernumber": sacrenum,
                                 "full_name": full_name,
                                 "all_functions": set()}
        persons[sacrenum]["all_functions"].add(label_raw)

    result = []
    for p in persons.values():
        sensitive = {f for f in p["all_functions"] if norm(f) in SACRE_SENSITIVE}
        if sensitive:
            result.append({**p, "sensitive_functions": sensitive})
    return result


# ── Parsing PKI (texte pipe-séparé) ───────────────────────────────────────────

def parse_pki(content: str) -> dict:
    """
    Accepte le contenu texte d'un export PKI psql pipe-séparé.
    Retourne dict norm(nom) → {name, groups: set}
    """
    m = re.search(
        r"Membres des Groupes\s*=+\s*(.*?)(?:\(\d+ rows\)|\Z)",
        content, re.DOTALL
    )
    if not m:
        raise ValueError("Section 'Membres des Groupes' introuvable dans le fichier PKI")

    persons = {}
    for line in m.group(1).splitlines():
        line  = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split("|")]
        if len(parts) < 2:
            continue
        if re.match(r"^[-+\s]+$", parts[0]) or parts[0].lower() in ("groupe", ""):
            continue
        group = parts[0]
        dn    = parts[1]
        if not group or not dn:
            continue
        cn_m = re.search(r"CN=([^,]+)", dn)
        if not cn_m:
            continue
        name = strip_operateur(cn_m.group(1))
        k    = norm(name)
        if k not in persons:
            persons[k] = {"name": name, "groups": set()}
        persons[k]["groups"].add(group)

    return persons


# ── Parsing KSTAMP (texte pipe-séparé) ────────────────────────────────────────

def parse_kstamp(content: str) -> list:
    """
    Accepte le contenu texte d'un export KSTAMP psql pipe-séparé.
    Colonnes : admin_dn | name | ws | ds | key | user | audit | admin  (t/f)
    """
    records = {}
    for line in content.splitlines():
        stripped = line.strip()
        if not stripped:
            continue
        parts = [p.strip() for p in stripped.split("|")]
        p0 = parts[0]

        if p0.startswith("+"):
            continue
        if re.match(r"^[-\s]+$", p0):
            continue
        if p0.lower() == "admin_dn":
            continue
        if re.match(r"^\(\d+ rows\)", stripped):
            continue
        if len(parts) < 7:
            continue

        name_raw = parts[1]
        if not name_raw or norm(name_raw) in KSTAMP_SKIP:
            continue

        name  = strip_operateur(name_raw)
        k     = norm(name)
        flags = {
            flag: (len(parts) > (2 + i) and parts[2 + i].lower() == "t")
            for i, flag in enumerate(_KSTAMP_FLAGS)
        }

        if any(flags[f] for f in ["workspace_management", "ds_management",
                                   "key_management", "user_management"]):
            profile = "admin"
        elif flags["audit_management"]:
            profile = "monitoring"
        else:
            profile = "admin_only"

        records[k] = {"name": name, "profile": profile, **flags}

    return list(records.values())


# ── Moteur d'analyse ──────────────────────────────────────────────────────────

def _disc(systeme, name, detail, status, entries):
    return {"systeme": systeme, "name": name, "detail": detail,
            "status": status, "roles_lir": entries}


def analyze_sacre(data, lir_direct, lir_wordset):
    results, disc = [], []
    for p in data:
        name    = p["full_name"]
        entries = lir_lookup(lir_direct, lir_wordset, name)
        status  = build_status(entries, [ROLE_ADMIN, ROLE_OPERATEUR], ["OSC"])
        results.append({"name": name, "functions": sorted(p["sensitive_functions"]),
                        "status": status, "roles_lir": entries})
        if status != "CONFORME":
            disc.append(_disc("SACRE", name,
                              ", ".join(sorted(p["sensitive_functions"])),
                              status, entries))
    return results, disc


def analyze_pki(pki_data, lir_direct, lir_wordset):
    results, disc = [], []
    for p in pki_data.values():
        name = p["name"]
        for group in sorted(p["groups"]):
            if group == "SERVER":
                results.append({"name": name, "group": group,
                                 "status": "N/A (compte serveur)", "roles_lir": []})
                continue
            allowed = [ROLE_ADMIN, ROLE_OPERATEUR] if group == "RUN" else [ROLE_ADMIN]
            entries = lir_lookup(lir_direct, lir_wordset, name)
            status  = build_status(entries, allowed, ["OSC"])
            results.append({"name": name, "group": group,
                             "status": status, "roles_lir": entries})
            if status != "CONFORME":
                disc.append(_disc("PKI", name, f"Groupe {group}", status, entries))
    return results, disc


def analyze_kstamp(data, lir_direct, lir_wordset, site):
    results, disc = [], []
    for p in data:
        name    = p["name"]
        profile = p["profile"]
        flags   = {f: p[f] for f in _KSTAMP_FLAGS}

        if profile == "admin_only":
            results.append({"name": name, "site": site, "profile": profile,
                             "status": "N/A (admin_management seul)",
                             "roles_lir": [], **flags})
            continue

        allowed = [ROLE_ADMIN] if profile == "admin" else [ROLE_ADMIN, ROLE_OPERATEUR]
        entries = lir_lookup(lir_direct, lir_wordset, name)
        status  = build_status(entries, allowed, ["OSH"])
        results.append({"name": name, "site": site, "profile": profile,
                         "status": status, "roles_lir": entries, **flags})
        if status != "CONFORME":
            disc.append(_disc(f"KSTAMP-{site}", name,
                              f"Profil {profile}", status, entries))
    return results, disc


# ── Génération rapport Excel (retourne des bytes) ─────────────────────────────

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED   = PatternFill("solid", fgColor="FFC7CE")
GREY  = PatternFill("solid", fgColor="D9D9D9")
HDR_F = PatternFill("solid", fgColor="4472C4")
HDR_T = Font(color="FFFFFF", bold=True)


def _status_fill(status: str) -> PatternFill:
    if status == "CONFORME":
        return GREEN
    if status.startswith("N/A"):
        return GREY
    return RED


def _write_header(ws, cols):
    ws.append(cols)
    for c in ws[ws.max_row]:
        c.fill = HDR_F
        c.font = HDR_T
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = ws["A2"]


def _auto_width(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 70)


def _roles_str(entries):
    return " | ".join(f"{r} / {d}" for r, d in entries) if entries else "—"


def generate_excel_bytes(control_date: str,
                         sacre_res, sacre_disc,
                         pki_res, pki_disc,
                         kstamp_sets: list) -> bytes:
    """
    kstamp_sets : liste de (site_label, results, discrepancies)
    Retourne le contenu du fichier Excel en bytes.
    """
    wb = openpyxl.Workbook()

    # SACRE
    ws = wb.active
    ws.title = "SACRE"
    _write_header(ws, ["Nom complet", "Fonctions sensibles détectées",
                        "Statut", "Rôles LIR (rôle / domaine)"])
    for r in sorted(sacre_res, key=lambda x: x["name"]):
        ws.append([r["name"], ", ".join(r["functions"]),
                   r["status"], _roles_str(r["roles_lir"])])
        for c in ws[ws.max_row]:
            c.fill = _status_fill(r["status"])
    _auto_width(ws)

    # PKI
    ws_pki = wb.create_sheet("PKI")
    _write_header(ws_pki, ["Nom complet", "Groupe PKI",
                             "Statut", "Rôles LIR (rôle / domaine)"])
    for r in sorted(pki_res, key=lambda x: (x["group"], x["name"])):
        ws_pki.append([r["name"], r["group"], r["status"],
                        _roles_str(r["roles_lir"])])
        for c in ws_pki[ws_pki.max_row]:
            c.fill = _status_fill(r["status"])
    _auto_width(ws_pki)

    # KSTAMP (une feuille par site)
    for site, res, _ in kstamp_sets:
        ws_k = wb.create_sheet(f"KSTAMP-{site}")
        _write_header(ws_k, ["Nom", "Site", "Profil",
                               "WS", "DS", "Key", "User", "Audit", "Admin",
                               "Statut", "Rôles LIR (rôle / domaine)"])
        for r in sorted(res, key=lambda x: x["name"]):
            f = lambda k: "oui" if r.get(k) else "non"
            ws_k.append([r["name"], r["site"], r["profile"],
                          f("workspace_management"), f("ds_management"),
                          f("key_management"),       f("user_management"),
                          f("audit_management"),     f("admin_management"),
                          r["status"], _roles_str(r.get("roles_lir", []))])
            for c in ws_k[ws_k.max_row]:
                c.fill = _status_fill(r["status"])
        _auto_width(ws_k)

    # ÉCARTS
    ws_e = wb.create_sheet("ÉCARTS")
    _write_header(ws_e, [
        "Système", "Nom", "Détail écart", "Statut",
        "Rôles LIR trouvés", "Objet ticket EasyVista",
        "Nomenclature", "Domaine", "Référentiel", "Sévérité", "Priorité",
    ])
    y, m = control_date[:4], control_date[5:7]
    all_kstamp_disc = [d for _, _, disc in kstamp_sets for d in disc]
    all_disc = sacre_disc + pki_disc + all_kstamp_disc
    for d in all_disc:
        ticket = f"[CONTROLE] {y}.{m}-Revue Droits Opérateurs - {d['name']}"
        ws_e.append([
            d["systeme"], d["name"], d["detail"], d["status"],
            _roles_str(d["roles_lir"]),
            ticket,
            "SSI/Incident de sécurité",
            "APPLICATIONS SITE CENTRAL",
            "SMSI",
            "Critique",
            "Haute",
        ])
        for c in ws_e[ws_e.max_row]:
            c.fill = RED
    _auto_width(ws_e)

    # RÉSUMÉ
    ws_s = wb.create_sheet("RÉSUMÉ", 0)
    ws_s["A1"] = f"Revue des droits opérateurs SACRE / PKI / KSTAMP  —  {control_date}"
    ws_s["A1"].font = Font(bold=True, size=13)
    ws_s.append([])
    _write_header(ws_s, ["Système", "Opérateurs contrôlés", "Conformes", "Écarts"])

    def _stats(results):
        ctrl = [r for r in results if not r["status"].startswith("N/A")]
        ok   = sum(1 for r in ctrl if r["status"] == "CONFORME")
        return len(ctrl), ok, len(ctrl) - ok

    for label, res in ([("SACRE", sacre_res), ("PKI", pki_res)]
                       + [(f"KSTAMP-{s}", r) for s, r, _ in kstamp_sets]):
        t, ok, ko = _stats(res)
        ws_s.append([label, t, ok, ko])
        ws_s[ws_s.max_row][3].fill = RED if ko > 0 else GREEN

    ws_s.append([])
    ws_s.append(["Total écarts", "", "", len(all_disc)])
    ws_s[ws_s.max_row][3].font = Font(bold=True)
    ws_s[ws_s.max_row][3].fill = RED if all_disc else GREEN
    _auto_width(ws_s)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Point d'entrée principal ──────────────────────────────────────────────────

def run_analysis(sacre_text: str, pki_text: str,
                 kstamp_texts: dict,
                 db_path: str,
                 control_date: str = None) -> dict:
    """
    kstamp_texts : {"MRS1": "<contenu>", "MRS2": "<contenu>", "CLY": "<contenu>"}
                   seuls les sites fournis sont analysés.
    db_path      : chemin vers baselir.db
    Retourne un dict JSON-sérialisable + la clé "excel_bytes" (bytes).
    """
    if not control_date:
        control_date = date.today().strftime("%Y-%m-%d")

    lir_direct, lir_wordset = load_lir_from_db(db_path)
    lir_count = len(lir_direct)

    sacre_data          = parse_sacre(sacre_text)
    sacre_res, sacre_disc = analyze_sacre(sacre_data, lir_direct, lir_wordset)

    pki_data          = parse_pki(pki_text)
    pki_res, pki_disc = analyze_pki(pki_data, lir_direct, lir_wordset)

    kstamp_sets = []
    for site in ("MRS1", "MRS2", "CLY"):
        if site in kstamp_texts and kstamp_texts[site]:
            data = parse_kstamp(kstamp_texts[site])
            res, disc = analyze_kstamp(data, lir_direct, lir_wordset, site)
            kstamp_sets.append((site, res, disc))

    all_disc = (sacre_disc + pki_disc
                + [d for _, _, disc in kstamp_sets for d in disc])

    excel_bytes = generate_excel_bytes(
        control_date, sacre_res, sacre_disc, pki_res, pki_disc, kstamp_sets
    )

    def _stats(results):
        ctrl = [r for r in results if not r["status"].startswith("N/A")]
        ok   = sum(1 for r in ctrl if r["status"] == "CONFORME")
        return {"total": len(ctrl), "conformes": ok, "ecarts": len(ctrl) - ok}

    return {
        "control_date":  control_date,
        "lir_count":     lir_count,
        "resume": {
            "SACRE": _stats(sacre_res),
            "PKI":   _stats(pki_res),
            **{f"KSTAMP-{s}": _stats(r) for s, r, _ in kstamp_sets},
            "total_ecarts": len(all_disc),
        },
        "sacre":         sacre_res,
        "pki":           pki_res,
        "kstamp_sets":   [
            {
                "site": site,
                "results": res,
                "discrepancies": disc,
            }
            for site, res, disc in kstamp_sets
        ],
        "ecarts":        all_disc,
        "excel_bytes":   excel_bytes,
    }
