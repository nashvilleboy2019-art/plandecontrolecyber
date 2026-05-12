from datetime import date
from fastapi import APIRouter, Request, Depends
from fastapi.responses import RedirectResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import io
from app.database import get_db
from app.models import Control, ControlResult, ControlType, Category, Perimetre, User
from app.utils import get_current_user, get_alert_status, current_period, compliance_color, periode_label
from app.templates_config import templates

router = APIRouter(tags=["dashboard"])


def _build_stats(db: Session, year: int = None):
    year = year or date.today().year
    controls = db.query(Control).filter(Control.archived == False).all()
    total = len(controls)

    # Latest result per control for the given year
    results_this_year = {}
    for c in controls:
        # Get the most recent result for this year
        r = next((res for res in c.results if res.annee == year and res.taux_conformite is not None), None)
        if r:
            results_this_year[c.id] = r

    compliant = sum(1 for r in results_this_year.values() if r.statut == "conforme")
    non_compliant = sum(1 for r in results_this_year.values() if r.statut == "non_conforme")
    pending = total - len(results_this_year)

    # Alerts
    overdue = [c for c in controls if get_alert_status(c, db) == "danger"]
    warnings = [c for c in controls if get_alert_status(c, db) == "warning"]

    # By type
    types = db.query(ControlType).filter(ControlType.active == True).order_by(ControlType.ordre).all()
    by_type = []
    for t in types:
        t_controls = [c for c in controls if c.type_id == t.id]
        t_results = [results_this_year[c.id] for c in t_controls if c.id in results_this_year]
        avg = (sum(r.taux_conformite for r in t_results) / len(t_results)) if t_results else None
        by_type.append({"label": t.label, "color": t.color, "count": len(t_controls), "avg": avg})

    # By category
    cats = db.query(Category).filter(Category.active == True).order_by(Category.ordre).all()
    by_category = []
    for cat in cats:
        c_controls = [c for c in controls if c.category_id == cat.id]
        c_results = [results_this_year[c.id] for c in c_controls if c.id in results_this_year]
        avg = (sum(r.taux_conformite for r in c_results) / len(c_results)) if c_results else None
        by_category.append({"label": cat.label, "count": len(c_controls), "avg": avg})

    # By perimeter
    perims = db.query(Perimetre).filter(Perimetre.active == True).order_by(Perimetre.ordre).all()
    by_perimetre = []
    for p in perims:
        p_controls = [c for c in controls if c.perimetre_id == p.id]
        p_results = [results_this_year[c.id] for c in p_controls if c.id in results_this_year]
        avg = (sum(r.taux_conformite for r in p_results) / len(p_results)) if p_results else None
        by_perimetre.append({"label": p.label, "count": len(p_controls), "avg": avg})

    # Monthly trend (all controls, average per month)
    monthly = []
    for m in range(1, 13):
        month_results = db.query(ControlResult).filter(
            ControlResult.annee == year,
            ControlResult.mois == m,
            ControlResult.taux_conformite.isnot(None),
        ).all()
        avg = (sum(r.taux_conformite for r in month_results) / len(month_results)) if month_results else None
        monthly.append({"mois": m, "avg": avg, "count": len(month_results)})

    # Unvalidated count
    unvalidated = db.query(ControlResult).filter(
        ControlResult.validated == False,
        ControlResult.taux_conformite.isnot(None),
    ).count()

    # Monthly heatmap by type (period start mapped to calendar month)
    def period_start_cal_month(freq, mois):
        if freq == "mensuel": return mois
        if freq == "bimestriel": return (mois - 1) * 2 + 1
        if freq == "trimestriel": return (mois - 1) * 3 + 1
        if freq == "semestriel": return 1 if mois == 1 else 7
        return 1  # annuel

    # Monthly completion count per category (réalisés / non effectués)
    # "Non effectué" = contrôle dont le mois m est le début de sa période
    # mais pour lequel aucun résultat avec taux n'existe pour cette période.
    def cal_month_to_period_local(freq, cal_month):
        if freq == "mensuel":    return cal_month
        if freq == "bimestriel": return (cal_month + 1) // 2
        if freq == "trimestriel": return (cal_month - 1) // 3 + 1
        if freq == "semestriel": return 1 if cal_month <= 6 else 2
        return 1

    # Index (control_id, annee, mois) → taux pour éviter les boucles O(n²)
    results_done_idx = {
        (r.control_id, r.annee, r.mois)
        for c in controls
        for r in c.results
        if r.taux_conformite is not None
    }

    cat_monthly_done = []
    cat_monthly_todo = []
    for cat in cats:
        cat_controls = [c for c in controls if c.category_id == cat.id]
        done_by_month = []
        todo_by_month = []
        for m in range(1, 13):
            done = 0
            todo = 0
            for ctrl in cat_controls:
                period_idx = cal_month_to_period_local(ctrl.frequence, m)
                # Ne compter le contrôle que sur le mois de début de sa période
                if period_start_cal_month(ctrl.frequence, period_idx) != m:
                    continue
                if (ctrl.id, year, period_idx) in results_done_idx:
                    done += 1
                else:
                    todo += 1
            done_by_month.append(done)
            todo_by_month.append(todo)
        cat_monthly_done.append({"label": cat.label, "data": done_by_month})
        cat_monthly_todo.append({"label": cat.label, "data": todo_by_month})

    monthly_by_type = []
    for t in types:
        t_controls = [c for c in controls if c.type_id == t.id]
        by_cal = {m: [] for m in range(1, 13)}
        for c in t_controls:
            for r in c.results:
                if r.annee == year and r.taux_conformite is not None:
                    cal_m = period_start_cal_month(c.frequence, r.mois)
                    if 1 <= cal_m <= 12:
                        by_cal[cal_m].append(r.taux_conformite)
        months_data = [
            round(sum(by_cal[m]) / len(by_cal[m]), 1) if by_cal[m] else None
            for m in range(1, 13)
        ]
        not_none = [v for v in months_data if v is not None]
        type_avg = round(sum(not_none) / len(not_none), 1) if not_none else None
        monthly_by_type.append({"label": t.label, "color": t.color, "months": months_data, "avg": type_avg})

    # Worst controls annual — top 10 by avg taux for the year
    worst_controls = []
    for c in controls:
        c_res = [r for r in c.results if r.annee == year and r.taux_conformite is not None]
        if not c_res:
            continue
        avg = round(sum(r.taux_conformite for r in c_res) / len(c_res), 1)
        non_conf = sum(1 for r in c_res if r.statut == "non_conforme")
        worst_controls.append({
            "id": c.id, "reference": c.reference, "libelle": c.libelle,
            "type_label": c.type.label if c.type else "—",
            "avg": avg, "non_conf": non_conf, "total": len(c_res),
            "taux_cible": c.taux_cible,
        })
    worst_controls.sort(key=lambda x: x["avg"])
    worst_controls = worst_controls[:10]

    # Worst controls this month — top 10 for current calendar month
    def period_covers_cal_month(freq, mois, cal_month):
        if freq == "mensuel": return mois == cal_month
        if freq == "bimestriel": return (mois - 1) * 2 + 1 <= cal_month <= mois * 2
        if freq == "trimestriel": return (mois - 1) * 3 + 1 <= cal_month <= mois * 3
        if freq == "semestriel": return (cal_month <= 6) if mois == 1 else (cal_month >= 7)
        return True  # annuel

    current_month = date.today().month
    worst_controls_month = []
    for c in controls:
        for r in c.results:
            if r.annee == year and r.taux_conformite is not None and period_covers_cal_month(c.frequence, r.mois, current_month):
                worst_controls_month.append({
                    "id": c.id, "reference": c.reference, "libelle": c.libelle,
                    "type_label": c.type.label if c.type else "—",
                    "taux": r.taux_conformite, "statut": r.statut,
                    "taux_cible": c.taux_cible,
                })
                break
    worst_controls_month.sort(key=lambda x: x["taux"])
    worst_controls_month = worst_controls_month[:10]

    # Indicators cockpit — grouped by domain
    import re as _re
    GROUP_LABELS = {
        "INC": "Incidents SECOPS",
        "VUM": "Vulnérabilités",
        "PSR": "Protection / CTI",
        "INF": "Infrastructure",
        "DEV": "Développement",
        "IGA": "Accès privilégiés",
        "CNF": "Conformité",
    }
    ind_type = next((t for t in types if t.label == "Indicateurs"), None)
    ind_controls = [c for c in controls if ind_type and c.type_id == ind_type.id]
    ind_items = []
    for c in ind_controls:
        c_res = sorted(
            [r for r in c.results if r.annee == year and r.taux_conformite is not None],
            key=lambda r: r.mois,
        )
        latest = c_res[-1] if c_res else None
        prev = c_res[-2] if len(c_res) >= 2 else None
        trend = None
        if latest and prev:
            d = latest.taux_conformite - prev.taux_conformite
            trend = "up" if d > 1 else ("down" if d < -1 else "stable")
        m = _re.search(r"IND-([A-Z]+)", c.reference)
        gc = m.group(1) if m else "AUTRE"
        ind_items.append({
            "id": c.id,
            "reference": c.reference,
            "libelle": c.libelle[:65],
            "frequence": c.frequence,
            "taux_cible": c.taux_cible,
            "latest": round(latest.taux_conformite, 1) if latest else None,
            "latest_label": latest.periode_label if latest else None,
            "statut": latest.statut if latest else "no_data",
            "trend": trend,
            "history": [round(r.taux_conformite, 1) for r in c_res],
            "history_labels": [r.periode_label for r in c_res],
            "group": gc,
            "group_label": GROUP_LABELS.get(gc, gc),
        })
    ind_items.sort(key=lambda x: (x["group"], x["reference"]))
    ind_with_data = [i for i in ind_items if i["latest"] is not None]
    ind_conf_count = sum(1 for i in ind_with_data if i["statut"] == "conforme")
    indicators_cockpit = {
        "items": ind_items,
        "score": round(ind_conf_count / len(ind_with_data) * 100, 1) if ind_with_data else None,
        "conf": ind_conf_count,
        "total": len(ind_items),
        "with_data": len(ind_with_data),
    }

    # By auditor — based on assigned_to_id on campaign results for the year
    by_auditor = []
    auditor_ids = [
        row[0] for row in
        db.query(ControlResult.assigned_to_id)
        .filter(ControlResult.annee == year, ControlResult.assigned_to_id.isnot(None))
        .distinct().all()
    ]
    for uid in auditor_ids:
        u = db.query(User).filter(User.id == uid).first()
        if not u:
            continue
        assigned = db.query(ControlResult).filter(
            ControlResult.annee == year,
            ControlResult.assigned_to_id == uid,
        ).all()
        done = [r for r in assigned if r.taux_conformite is not None]
        avg = round(sum(r.taux_conformite for r in done) / len(done), 1) if done else None
        by_auditor.append({
            "label": u.nom_complet or u.username,
            "count": len(assigned),
            "done": len(done),
            "avg": avg,
        })
    by_auditor.sort(key=lambda x: x["count"], reverse=True)

    return {
        "year": year,
        "total": total,
        "compliant": compliant,
        "non_compliant": non_compliant,
        "pending": pending,
        "overdue": overdue,
        "warnings": warnings,
        "by_type": by_type,
        "by_category": by_category,
        "by_perimetre": by_perimetre,
        "monthly": monthly,
        "unvalidated": unvalidated,
        "cat_monthly_done": cat_monthly_done,
        "cat_monthly_todo": cat_monthly_todo,
        "by_auditor": by_auditor,
        "monthly_by_type": monthly_by_type,
        "worst_controls": worst_controls,
        "worst_controls_month": worst_controls_month,
        "current_month": current_month,
        "indicators_cockpit": indicators_cockpit,
        "compliance_rate": round(
            sum(r.taux_conformite for r in results_this_year.values()) / len(results_this_year), 1
        ) if results_this_year else None,
    }


@router.get("/dashboard")
async def dashboard(request: Request, db: Session = Depends(get_db), year: int = None):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)

    today_year = date.today().year
    year = year or today_year
    stats = _build_stats(db, year)

    # Only show years that have at least one result, plus the current year
    years_with_data = {
        r.annee for r in db.query(ControlResult.annee).distinct().all()
    }
    years_with_data.add(today_year)
    available_years = sorted(years_with_data)

    incidents_en_cours = (
        db.query(ControlResult)
        .join(Control, ControlResult.control_id == Control.id)
        .filter(
            ControlResult.statut == "incident_en_cours",
            ControlResult.validated == False,
            Control.archived == False,
        )
        .order_by(ControlResult.updated_at.desc())
        .all()
    )
    for inc in incidents_en_cours:
        inc._period_label = periode_label(inc.control.frequence, inc.annee, inc.mois)

    return templates.TemplateResponse(request, "dashboard.html", {
        "request": request, "user": user,
        "stats": stats,
        "available_years": available_years,
        "current_year": year,
        "incidents_en_cours": incidents_en_cours,
    })


@router.get("/dashboard/export")
async def export_dashboard(request: Request, db: Session = Depends(get_db), year: int = None):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login", status_code=302)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.utils import MOIS_LABELS

    year = year or date.today().year
    stats = _build_stats(db, year)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Dashboard {year}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill_blue = PatternFill("solid", fgColor="1E3A5F")
    header_fill_gray = PatternFill("solid", fgColor="475569")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    orange_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill_for(rate):
        if rate is None:
            return None
        if rate >= 90:
            return green_fill
        if rate >= 70:
            return orange_fill
        return red_fill

    row = 1

    # Title
    ws.merge_cells(f"A{row}:H{row}")
    title_cell = ws[f"A{row}"]
    title_cell.value = f"Tableau de Bord – Plan de Contrôle Cyber – {year}"
    title_cell.font = Font(bold=True, size=14, color="1E3A5F")
    title_cell.alignment = center
    row += 2

    # Summary
    summary_headers = ["Total contrôles", "Conformes", "Non conformes", "Sans résultat",
                       "En retard", "Taux global"]
    for i, h in enumerate(summary_headers, 1):
        c = ws.cell(row, i, h)
        c.font = header_font; c.fill = header_fill_blue; c.alignment = center; c.border = border
    row += 1
    summary_vals = [
        stats["total"], stats["compliant"], stats["non_compliant"],
        stats["pending"], len(stats["overdue"]),
        f"{stats['compliance_rate']}%" if stats["compliance_rate"] is not None else "N/A",
    ]
    for i, v in enumerate(summary_vals, 1):
        c = ws.cell(row, i, v)
        c.alignment = center; c.border = border
    row += 2

    # By type
    ws.cell(row, 1, "Conformité par thématique").font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for h, col in [("Thématique", 1), ("Nb contrôles", 2), ("Taux moyen (%)", 3)]:
        c = ws.cell(row, col, h)
        c.font = header_font; c.fill = header_fill_gray; c.alignment = center; c.border = border
    row += 1
    for t in stats["by_type"]:
        ws.cell(row, 1, t["label"]).border = border
        ws.cell(row, 2, t["count"]).alignment = center; ws.cell(row, 2).border = border
        val = round(t["avg"], 1) if t["avg"] is not None else "N/A"
        c = ws.cell(row, 3, val)
        c.alignment = center; c.border = border
        if t["avg"] is not None:
            c.fill = fill_for(t["avg"])
        row += 1
    row += 1

    # By category
    ws.cell(row, 1, "Conformité par catégorie").font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for h, col in [("Catégorie", 1), ("Nb contrôles", 2), ("Taux moyen (%)", 3)]:
        c = ws.cell(row, col, h)
        c.font = header_font; c.fill = header_fill_gray; c.alignment = center; c.border = border
    row += 1
    for cat in stats["by_category"]:
        ws.cell(row, 1, cat["label"]).border = border
        ws.cell(row, 2, cat["count"]).alignment = center; ws.cell(row, 2).border = border
        val = round(cat["avg"], 1) if cat["avg"] is not None else "N/A"
        c = ws.cell(row, 3, val)
        c.alignment = center; c.border = border
        if cat["avg"] is not None:
            c.fill = fill_for(cat["avg"])
        row += 1
    row += 1

    # Monthly trend
    ws.cell(row, 1, "Tendance mensuelle").font = Font(bold=True, size=12, color="1E3A5F")
    row += 1
    for m_data in stats["monthly"]:
        c = ws.cell(row, m_data["mois"], MOIS_LABELS[m_data["mois"]])
        c.font = header_font; c.fill = header_fill_gray; c.alignment = center; c.border = border
    row += 1
    for m_data in stats["monthly"]:
        val = round(m_data["avg"], 1) if m_data["avg"] is not None else "N/A"
        c = ws.cell(row, m_data["mois"], val)
        c.alignment = center; c.border = border
        if m_data["avg"] is not None:
            c.fill = fill_for(m_data["avg"])
    row += 2

    # Overdue controls sheet
    ws2 = wb.create_sheet("Contrôles en retard")
    for h, col in [("Référence", 1), ("Libellé", 2), ("Fréquence", 3), ("Responsable", 4)]:
        c = ws2.cell(1, col, h)
        c.font = header_font; c.fill = header_fill_blue; c.alignment = center; c.border = border
    for i, ctrl in enumerate(stats["overdue"], 2):
        ws2.cell(i, 1, ctrl.reference).border = border
        ws2.cell(i, 2, ctrl.libelle).border = border
        ws2.cell(i, 3, ctrl.frequence).alignment = center; ws2.cell(i, 3).border = border
        ws2.cell(i, 4, ctrl.responsable.nom_complet if ctrl.responsable else "—").border = border
    ws2.column_dimensions["B"].width = 50

    # Detail sheet
    ws3 = wb.create_sheet("Détail contrôles")
    headers3 = ["Référence", "Libellé", "Thématique", "Catégorie", "Périmètre",
                "Fréquence", "Responsable", "Taux cible"]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(1, col, h)
        c.font = header_font; c.fill = header_fill_blue; c.alignment = center; c.border = border
    controls = db.query(Control).filter(Control.archived == False).order_by(Control.reference).all()
    for i, ctrl in enumerate(controls, 2):
        ws3.cell(i, 1, ctrl.reference).border = border
        ws3.cell(i, 2, ctrl.libelle).border = border
        ws3.cell(i, 3, ctrl.type.label if ctrl.type else "—").border = border
        ws3.cell(i, 4, ctrl.category.label if ctrl.category else "—").border = border
        ws3.cell(i, 5, ctrl.perimetre.label if ctrl.perimetre else "—").border = border
        ws3.cell(i, 6, ctrl.frequence).alignment = center; ws3.cell(i, 6).border = border
        ws3.cell(i, 7, ctrl.responsable.nom_complet if ctrl.responsable else "—").border = border
        ws3.cell(i, 8, f"{ctrl.taux_cible}%").alignment = center; ws3.cell(i, 8).border = border

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dashboard_pdc_{year}.xlsx"}
    )
