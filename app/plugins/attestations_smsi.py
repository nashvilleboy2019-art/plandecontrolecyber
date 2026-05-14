"""
Plugin : Taux d'attestations conformes — Rôles de confiance SMSI (DSO-ATT-01)
Interroge BaseLIR, filtre domaine SMSI, vérifie attestation_expiree.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SLUG            = "attestations_smsi"
FORM_TEMPLATE   = "plugins/attestations_smsi/form.html"
RESULT_TEMPLATE = "plugins/attestations_smsi/resultats.html"

STATUS_CONFORME  = "CONFORME"
STATUS_EXPIREE   = "ATTESTATION EXPIRÉE"
STATUS_MANQUANTE = "ATTESTATION MANQUANTE"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
HDR_F  = PatternFill("solid", fgColor="4472C4")
HDR_T  = Font(color="FFFFFF", bold=True)


def _statut(hab: dict) -> str:
    exp = hab.get("attestation_expiree")
    if exp is None:
        return STATUS_MANQUANTE
    return STATUS_EXPIREE if exp else STATUS_CONFORME


def _fetch_smsi(lir_url: str, lir_key: str) -> list:
    import requests as _req
    headers  = {"X-API-Key": lir_key}
    page, pp = 1, 200
    items    = []
    while True:
        resp = _req.get(
            f"{lir_url.rstrip('/')}/api/v1/habilitations",
            headers=headers,
            params={"statut_id": 1, "per_page": pp, "page": page},
            timeout=15,
        )
        resp.raise_for_status()
        data  = resp.json()
        batch = data.get("items", data) if isinstance(data, dict) else data
        for h in batch:
            if (h.get("domaine") or "").strip().upper() == "SMSI":
                items.append(h)
        if len(batch) < pp:
            break
        page += 1
    return items


async def execute(form, config: dict, lir_url: str, lir_key: str, control_date: str) -> dict:
    if not lir_key:
        raise ValueError("Clé API BaseLIR non configurée (Admin → Plugins → Connexion BaseLIR).")

    habs = _fetch_smsi(lir_url, lir_key)
    if not habs:
        raise ValueError("Aucune habilitation active avec domaine SMSI trouvée dans BaseLIR.")

    results, ecarts = [], []
    for h in sorted(habs, key=lambda x: (x.get("nom_prenom") or "").lower()):
        status = _statut(h)
        entry  = {
            "name":                 h.get("nom_prenom", ""),
            "role":                 h.get("role") or "—",
            "date_attestation":     h.get("date_attestation"),
            "attestation_filename": h.get("attestation_filename"),
            "status":               status,
        }
        results.append(entry)
        if status != STATUS_CONFORME:
            ecarts.append({
                "systeme": "ATT",
                "name":    entry["name"],
                "detail":  f"{entry['role']} — {status}",
                "status":  status,
                "roles_lir": [],
            })

    total     = len(results)
    conformes = sum(1 for r in results if r["status"] == STATUS_CONFORME)
    n_ecarts  = total - conformes

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Attestations SMSI"
    ws.append(["Nom / Prénom", "Rôle", "Date d'attestation", "Statut"])
    for c in ws[ws.max_row]:
        c.fill = HDR_F; c.font = HDR_T
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    fill_map = {STATUS_CONFORME: GREEN, STATUS_EXPIREE: ORANGE, STATUS_MANQUANTE: RED}
    for r in results:
        ws.append([r["name"], r["role"], r["date_attestation"] or "—", r["status"]])
        fill = fill_map.get(r["status"], RED)
        for c in ws[ws.max_row]:
            c.fill = fill

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 60)

    buf = io.BytesIO()
    wb.save(buf)

    return {
        "control_date": control_date,
        "lir_count":    len(habs),
        "resume": {
            "ATT": {"total": total, "conformes": conformes, "ecarts": n_ecarts},
            "total_ecarts": n_ecarts,
        },
        "results": results,
        "ecarts":  ecarts,
        "excel_bytes": buf.getvalue(),
    }


def compute_taux(result: dict) -> float:
    att    = result.get("resume", {}).get("ATT", {})
    total  = att.get("total", 0)
    ecarts = att.get("ecarts", 0)
    if total == 0:
        return 100.0
    return round((total - ecarts) / total * 100, 1)


def build_commentaire(result: dict) -> str:
    att    = result.get("resume", {}).get("ATT", {})
    nb     = att.get("ecarts", 0)
    total  = att.get("total", 0)
    date_s = result.get("control_date", "")
    if nb == 0:
        return f"Attestations SMSI du {date_s} : {total} rôle(s) de confiance contrôlé(s), aucun écart."
    return (f"Attestations SMSI du {date_s} : {nb} écart(s) sur {total} rôle(s) de confiance "
            f"(attestation manquante ou expirée). Voir rapport Excel.")
