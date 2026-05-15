"""
Plugin : Revue des Droits Opérateurs – BASTION SIN (DSO-LOG-SIN)
Croise l'export CSV de la console BASTION SI Notaire avec BaseLIR.
"""

import io, csv, unicodedata, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SLUG            = "revue_droits_bastion_sin"
FORM_TEMPLATE   = "plugins/revue_droits_bastion_sin/form.html"
RESULT_TEMPLATE = "plugins/revue_droits_bastion_sin/resultats.html"

STATUS_CONFORME       = "CONFORME"
STATUS_NON_CONFORME   = "NON CONFORME"
STATUS_NON_TROUVE     = "NON TROUVÉ LIR"
STATUS_SERVICE        = "N/A - COMPTE DE SERVICE"
STATUS_BRIS_DE_GLACE  = "N/A - BRIS DE GLACE"
STATUS_PROFIL_INCONNU = "PROFIL INCONNU"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
GRAY   = PatternFill("solid", fgColor="EDEDED")
HDR_F  = PatternFill("solid", fgColor="4472C4")
HDR_T  = Font(color="FFFFFF", bold=True)

CONFORMITY_RULES = {
    "product_administrator": {
        "roles":    ["ingénieur / administrateur système"],
        "domaines": ["infrastructure"],
        "services": ["sisn"],
    },
    "auditor": {
        "roles":    ["responsable de la sécurité"],
        "domaines": ["groupe adsn"],
        "services": ["secops/csirt", "csirt"],
    },
    "infrastructure_administrator": {
        "roles":    ["ingénieur / administrateur système"],
        "domaines": ["infrastructure"],
        "services": ["siei"],
    },
}

_SERVICE_KEYWORDS = {"administrator", "admin", "service", "system", "super", "svc", "bot"}


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().upper()


def _parse_csv(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    lines = [l for l in text.splitlines() if l.strip()]

    start = 0
    if lines and lines[0].lstrip().startswith("#wab"):
        start = 1

    sample = lines[start] if start < len(lines) else ""
    delim = ";" if sample.count(";") > sample.count(",") else ","

    reader = csv.DictReader(lines[start:], delimiter=delim)
    rows = []
    for row in reader:
        normalized = {(k or "").lstrip("#").strip(): (v or "").strip()
                      for k, v in row.items()}
        rows.append(normalized)
    return rows


def _get_col(row: dict, *candidates: str) -> str:
    for candidate in candidates:
        for k, v in row.items():
            if k.strip().lower() == candidate.lower():
                return v
    return ""


def _is_service_account(username: str, utilisateur: str) -> bool:
    if not utilisateur:
        return True
    if utilisateur.lower() == username.lower():
        return True
    if "." in utilisateur or "_" in utilisateur:
        return True
    if set(utilisateur.lower().split()) & _SERVICE_KEYWORDS:
        return True
    return False


def _parse_nom_prenom(utilisateur: str) -> tuple[str, str]:
    parts = utilisateur.strip().split()
    if len(parts) < 2:
        return (utilisateur, "")
    return (parts[-1], " ".join(parts[:-1]))


def _fetch_lir(lir_url: str, lir_key: str) -> dict:
    import requests as _req
    base = lir_url.rstrip("/")
    headers = {"X-API-Key": lir_key}
    pp, page = 200, 1
    lir: dict[str, list] = {}
    while True:
        resp = _req.get(
            f"{base}/api/v1/habilitations",
            headers=headers,
            params={"per_page": pp, "page": page},
            timeout=15,
        )
        resp.raise_for_status()
        data  = resp.json()
        batch = data.get("items", data) if isinstance(data, dict) else data
        for h in batch:
            key = _norm(h.get("nom_prenom", ""))
            if not key:
                continue
            lir.setdefault(key, []).append({
                "role":      h.get("role", "") or "",
                "domaine":   h.get("domaine", "") or "",
                "direction": h.get("direction", "") or "",
                "service":   h.get("service", "") or "",
            })
        if len(batch) < pp:
            break
        page += 1
    return lir


def _check_conformity(profile: str, lir_entries: list[dict],
                      rules_map: dict | None = None) -> str:
    rules_map = rules_map or CONFORMITY_RULES
    rules = rules_map.get(profile.strip().lower())
    if rules is None:
        return STATUS_PROFIL_INCONNU
    for entry in lir_entries:
        role_v    = _norm(entry["role"])
        domaine_v = _norm(entry["domaine"])
        service_v = _norm(entry["service"])
        # If the LIR entry has no value for a dimension, skip that dimension's check.
        role_ok    = (not role_v)    or any(_norm(r) in role_v    for r in rules["roles"])
        domaine_ok = (not domaine_v) or any(_norm(d) in domaine_v for d in rules["domaines"])
        service_ok = (not service_v) or any(_norm(s) in service_v for s in rules["services"])
        if role_ok and domaine_ok and service_ok:
            return STATUS_CONFORME
    return STATUS_NON_CONFORME


async def execute(form, config: dict, lir_url: str, lir_key: str, control_date: str) -> dict:
    if not lir_key:
        raise ValueError("Clé API BaseLIR non configurée (Admin → Plugins → Connexion BaseLIR).")

    bastion_file = form.get("bastion_file")
    if not bastion_file or not getattr(bastion_file, "filename", None):
        raise ValueError("Le fichier export BASTION SIN est requis.")

    file_bytes = await bastion_file.read()
    rows = _parse_csv(file_bytes)
    if not rows:
        raise ValueError("Le fichier CSV est vide ou illisible.")

    lir = _fetch_lir(lir_url, lir_key)
    lir_count = sum(len(v) for v in lir.values())

    custom_rules = config.get("plugin_rules") or {}
    effective_rules = {**CONFORMITY_RULES, **custom_rules}

    results, ecarts = [], []

    for row in rows:
        username    = _get_col(row, "Compte_utilisateur", "Username")
        utilisateur = _get_col(row, "utilisateur", "Full name", "Full_name")
        profile     = _get_col(row, "profile_name", "Profile")

        if not username:
            continue

        # Bris de glace → not evaluated
        if profile.lower() == "brisdeglace":
            results.append({
                "username":    username,
                "name":        utilisateur or username,
                "profile":     profile,
                "type_compte": "Bris de glace",
                "nom_tech":    "",
                "role_lir":    "",
                "domaine_lir": "",
                "service_lir": "",
                "status":      STATUS_BRIS_DE_GLACE,
                "detail":      "Compte bris de glace — non évalué",
            })
            continue

        if _is_service_account(username, utilisateur):
            results.append({
                "username":    username,
                "name":        utilisateur or username,
                "profile":     profile,
                "type_compte": "Compte de service",
                "nom_tech":    "",
                "role_lir":    "",
                "domaine_lir": "",
                "service_lir": "",
                "status":      STATUS_SERVICE,
                "detail":      "Compte de service — non évalué",
            })
            continue

        nom, prenom = _parse_nom_prenom(utilisateur)
        nom_tech    = f"{_norm(nom)} {_norm(prenom)}".strip()
        lir_entries = lir.get(nom_tech, [])

        if not lir_entries:
            status      = STATUS_NON_TROUVE
            role_lir = domaine_lir = service_lir = ""
            detail = f"Profil: {profile} — NOM TECHNIQUE introuvable dans BaseLIR ({nom_tech})"
        else:
            status      = _check_conformity(profile, lir_entries, effective_rules)
            best        = lir_entries[0]
            role_lir    = best["role"]
            domaine_lir = best["domaine"]
            service_lir = best["service"]
            if status == STATUS_CONFORME:
                detail = f"Profil: {profile} — {role_lir} / {domaine_lir} / {service_lir}"
            elif status == STATUS_PROFIL_INCONNU:
                detail = f"Profil applicatif absent de la matrice de contrôle : {profile}"
            else:
                rules   = effective_rules.get(profile.strip().lower(), {})
                exp_r   = ", ".join(rules.get("roles", []))
                exp_s   = ", ".join(rules.get("services", []))
                detail  = (f"Profil: {profile} — LIR: {role_lir} / {domaine_lir} / {service_lir} "
                           f"— Attendu rôle: {exp_r} / service: {exp_s}")

        entry = {
            "username":    username,
            "name":        utilisateur,
            "profile":     profile,
            "type_compte": "Nominatif",
            "nom_tech":    nom_tech,
            "role_lir":    role_lir,
            "domaine_lir": domaine_lir,
            "service_lir": service_lir,
            "status":      status,
            "detail":      detail,
        }
        results.append(entry)
        if status not in (STATUS_CONFORME, STATUS_SERVICE):
            ecarts.append({
                "systeme":   "BASTION SIN",
                "name":      utilisateur,
                "username":  username,
                "profile":   profile,
                "detail":    detail,
                "status":    status,
                "roles_lir": [],
            })

    nominatifs = [r for r in results if r["type_compte"] == "Nominatif"]
    total     = len(nominatifs)
    conformes = sum(1 for r in nominatifs if r["status"] == STATUS_CONFORME)
    n_ecarts  = total - conformes

    # Export Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "BASTION SIN"
    ws.append(["Login", "Nom complet", "Profil applicatif", "Type de compte",
               "NOM TECHNIQUE", "Rôle LIR", "Domaine LIR", "Service LIR", "Conformité"])
    for c in ws[ws.max_row]:
        c.fill = HDR_F; c.font = HDR_T
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    fill_map = {
        STATUS_CONFORME:       GREEN,
        STATUS_NON_CONFORME:   RED,
        STATUS_NON_TROUVE:     ORANGE,
        STATUS_SERVICE:        GRAY,
        STATUS_BRIS_DE_GLACE:  GRAY,
        STATUS_PROFIL_INCONNU: RED,
    }
    for r in results:
        ws.append([r["username"], r["name"], r["profile"], r["type_compte"],
                   r["nom_tech"], r["role_lir"], r["domaine_lir"],
                   r["service_lir"], r["status"]])
        for c in ws[ws.max_row]:
            c.fill = fill_map.get(r["status"], RED)

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 50)

    buf = io.BytesIO()
    wb.save(buf)

    return {
        "control_date":   control_date,
        "lir_count":      lir_count,
        "total_accounts": len(rows),
        "resume": {
            "BASTION SIN": {"total": total, "conformes": conformes, "ecarts": n_ecarts},
            "total_ecarts": n_ecarts,
        },
        "results": results,
        "ecarts":  ecarts,
        "excel_bytes": buf.getvalue(),
    }


def compute_taux(result: dict) -> float:
    b      = result.get("resume", {}).get("BASTION SIN", {})
    total  = b.get("total", 0)
    ecarts = b.get("ecarts", 0)
    if total == 0:
        return 100.0
    return round((total - ecarts) / total * 100, 1)


def build_commentaire(result: dict) -> str:
    b      = result.get("resume", {}).get("BASTION SIN", {})
    nb     = b.get("ecarts", 0)
    total  = b.get("total", 0)
    date_s = result.get("control_date", "")
    if nb == 0:
        return (f"Contrôle BASTION SIN du {date_s} : {total} compte(s) nominatif(s) "
                f"contrôlé(s), tous conformes.")
    return (f"Contrôle BASTION SIN du {date_s} : {nb} écart(s) sur {total} "
            f"compte(s) nominatif(s). Voir rapport Excel pour le détail.")
