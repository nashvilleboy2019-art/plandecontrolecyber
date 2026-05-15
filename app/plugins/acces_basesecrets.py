"""
Plugin : Contrôle Accès Base des Secrets (DSO-ACC-01)
Vérifie que les logs de connexion BaseSECRETS ne contiennent que des
utilisateurs enregistrés (liste locale autorisée).
"""

import io, calendar
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SLUG            = "acces_basesecrets"
FORM_TEMPLATE   = "plugins/acces_basesecrets/form.html"
RESULT_TEMPLATE = "plugins/acces_basesecrets/resultats.html"

STATUS_CONFORME     = "CONFORME"
STATUS_NON_AUTORISE = "NON AUTORISÉ"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
HDR_F  = PatternFill("solid", fgColor="4472C4")
HDR_T  = Font(color="FFFFFF", bold=True)


def _fetch_secrets_users(secrets_url: str, secrets_key: str) -> list:
    import requests as _req
    resp = _req.get(
        f"{secrets_url.rstrip('/')}/api/v1/users",
        headers={"X-API-Key": secrets_key},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()


def _fetch_connections(secrets_url: str, secrets_key: str,
                       from_date: str, to_date: str) -> list:
    import requests as _req
    resp = _req.get(
        f"{secrets_url.rstrip('/')}/api/v1/connections",
        headers={"X-API-Key": secrets_key},
        params={"from_date": from_date, "to_date": to_date},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json()


async def execute(form, config: dict, lir_url: str, lir_key: str, control_date: str) -> dict:
    secrets_url = config.get("secrets_url", "")
    secrets_key = config.get("secrets_key", "")

    if not secrets_key:
        raise ValueError("Clé API BaseSECRETS non configurée (Admin → Plugins → Connexion BaseSECRETS).")

    today = date.today()
    annee    = int(form.get("annee") or today.year)
    mois     = int(form.get("mois")  or today.month)
    last_day = calendar.monthrange(annee, mois)[1]
    from_date = f"{annee:04d}-{mois:02d}-01"
    to_date   = f"{annee:04d}-{mois:02d}-{last_day:02d}"

    users = _fetch_secrets_users(secrets_url, secrets_key)
    registered = {u["username"] for u in users}

    connections = _fetch_connections(secrets_url, secrets_key, from_date, to_date)
    connected_usernames = {c["username"] for c in connections}

    from collections import Counter
    conn_by_user = Counter(c["username"] for c in connections)

    results = []
    for u in sorted(users, key=lambda x: (x.get("last_name") or x.get("username") or "").lower()):
        last  = (u.get("last_name") or "").strip()
        first = (u.get("first_name") or "").strip()
        nom_prenom = (u.get("nom_prenom")
                      or (f"{last.upper()} {first.capitalize()}".strip() if last else "")
                      or u.get("username", ""))
        results.append({
            "name":       nom_prenom,
            "username":   u.get("username", ""),
            "role":       u.get("role", ""),
            "created_at": (u.get("created_at") or "")[:10],
            "status":     STATUS_CONFORME,
        })

    ecarts = []
    for username in sorted(connected_usernames - registered):
        entry = {
            "name":       username,
            "username":   username,
            "role":       "—",
            "created_at": "",
            "status":     STATUS_NON_AUTORISE,
        }
        results.append(entry)
        ecarts.append(entry.copy())

    total     = len(results)
    conformes = len(users)
    n_ecarts  = len(ecarts)

    wb = Workbook()
    ws = wb.active
    ws.title = "Accès BaseSECRETS"
    ws.append(["Nom / Prénom", "Login", "Rôle", "Créé le", "Statut"])
    for c in ws[ws.max_row]:
        c.fill = HDR_F; c.font = HDR_T
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    fill_map = {STATUS_CONFORME: GREEN, STATUS_NON_AUTORISE: RED}
    for r in results:
        ws.append([r["name"], r["username"], r["role"], r["created_at"], r["status"]])
        for c in ws[ws.max_row]:
            c.fill = fill_map.get(r["status"], RED)

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 50)

    buf = io.BytesIO()
    wb.save(buf)

    connections_sorted = sorted(connections, key=lambda c: c.get("timestamp", ""), reverse=True)

    return {
        "control_date":  control_date,
        "secrets_count": len(users),
        "conn_count":    len(connections),
        "conn_by_user":  dict(conn_by_user),
        "connections":   connections_sorted,
        "resume": {
            "ACC": {"total": total, "conformes": conformes, "ecarts": n_ecarts},
            "total_ecarts": n_ecarts,
        },
        "results": results,
        "ecarts":  ecarts,
        "excel_bytes": buf.getvalue(),
    }


def compute_taux(result: dict) -> float:
    acc    = result.get("resume", {}).get("ACC", {})
    total  = acc.get("total", 0)
    ecarts = acc.get("ecarts", 0)
    if total == 0:
        return 100.0
    return round((total - ecarts) / total * 100, 1)


def build_commentaire(result: dict) -> str:
    acc    = result.get("resume", {}).get("ACC", {})
    nb     = acc.get("ecarts", 0)
    total  = result.get("secrets_count", 0)
    conn   = result.get("conn_count", 0)
    date_s = result.get("control_date", "")
    if nb == 0:
        return (f"Accès BaseSECRETS du {date_s} : {total} compte(s) enregistré(s), "
                f"{conn} connexion(s) vérifiée(s) — toutes provenant d'utilisateurs autorisés.")
    return (f"Accès BaseSECRETS du {date_s} : {nb} connexion(s) non autorisée(s) détectée(s) "
            f"(utilisateur(s) absent(s) de la liste locale BaseSECRETS). Voir rapport Excel.")
