"""
Plugin : Taux d'attestations conformes — Sensibilisations sécurité SMSI (DSO-ATT-02)
Interroge BaseLIR, filtre statuts Sous-traitant / Centre de service,
vérifie sensibilisation_expiree.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SLUG            = "sensibilisations_smsi"
FORM_TEMPLATE   = "plugins/sensibilisations_smsi/form.html"
RESULT_TEMPLATE = "plugins/sensibilisations_smsi/resultats.html"

STATUS_CONFORME  = "CONFORME"
STATUS_EXPIREE   = "SENSIBILISATION EXPIRÉE"
STATUS_MANQUANTE = "SENSIBILISATION MANQUANTE"

GREEN  = PatternFill("solid", fgColor="C6EFCE")
ORANGE = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
HDR_F  = PatternFill("solid", fgColor="4472C4")
HDR_T  = Font(color="FFFFFF", bold=True)

STATUT_LABELS = ["Sous-traitant", "Centre de service"]


def _statut(hab: dict) -> str:
    exp = hab.get("sensibilisation_expiree")
    if exp is None:
        return STATUS_MANQUANTE
    return STATUS_EXPIREE if exp else STATUS_CONFORME


def _get_statut_ids(base_url: str, headers: dict) -> list[int]:
    import requests as _req
    resp = _req.get(f"{base_url}/api/v1/referentiels", headers=headers, timeout=10)
    resp.raise_for_status()
    statuts = resp.json().get("statuts", [])
    found = [
        s["id"] for s in statuts
        if any(lbl.lower() in (s.get("label") or "").strip().lower()
               for lbl in STATUT_LABELS)
    ]
    if not found:
        labels_dispo = [s.get("label", "") for s in statuts]
        raise ValueError(
            f"Aucun statut « Sous-traitant » ou « Centre de service » trouvé dans BaseLIR. "
            f"Statuts disponibles : {', '.join(labels_dispo) or '(aucun)'}."
        )
    return found


def _fetch_habilitations(lir_url: str, lir_key: str) -> list:
    import requests as _req
    base    = lir_url.rstrip("/")
    headers = {"X-API-Key": lir_key}

    statut_ids = _get_statut_ids(base, headers)

    pp    = 200
    seen  = set()
    items = []
    for statut_id in statut_ids:
        page = 1
        while True:
            resp = _req.get(
                f"{base}/api/v1/habilitations",
                headers=headers,
                params={"statut_id": statut_id, "per_page": pp, "page": page},
                timeout=15,
            )
            resp.raise_for_status()
            data  = resp.json()
            batch = data.get("items", data) if isinstance(data, dict) else data
            for h in batch:
                uid = h.get("id") or h.get("nom_prenom", "")
                if uid not in seen:
                    seen.add(uid)
                    items.append(h)
            if len(batch) < pp:
                break
            page += 1
    return items


_STATUS_SEVERITY = {STATUS_MANQUANTE: 2, STATUS_EXPIREE: 1, STATUS_CONFORME: 0}


def _group_ecarts(raw: list) -> list:
    """Regroupe les écarts par personne : fusionne les domaines, garde le statut le plus grave."""
    merged: dict = {}
    order: list  = []
    for e in raw:
        key = e["name"]
        if key not in merged:
            merged[key] = e.copy()
            merged[key]["_domaines"]  = [e["domaine"]] if e.get("domaine", "—") != "—" else []
            merged[key]["_hab_count"] = 1
            order.append(key)
        else:
            dom = e.get("domaine", "—")
            if dom != "—" and dom not in merged[key]["_domaines"]:
                merged[key]["_domaines"].append(dom)
            merged[key]["_hab_count"] += 1
            if _STATUS_SEVERITY.get(e["status"], 0) > _STATUS_SEVERITY.get(merged[key]["status"], 0):
                merged[key]["status"] = e["status"]
    result = []
    for key in order:
        e = merged[key]
        e["domaine"] = " / ".join(e["_domaines"]) if e["_domaines"] else "—"
        result.append(e)
    return result


async def execute(form, config: dict, lir_url: str, lir_key: str, control_date: str) -> dict:
    if not lir_key:
        raise ValueError("Clé API BaseLIR non configurée (Admin → Plugins → Connexion BaseLIR).")

    habs = _fetch_habilitations(lir_url, lir_key)
    if not habs:
        raise ValueError("Aucune habilitation active (Sous-traitant / Centre de service) trouvée dans BaseLIR.")

    results, ecarts = [], []
    for h in sorted(habs, key=lambda x: (x.get("nom_prenom") or "").lower()):
        status = _statut(h)
        entry  = {
            "name":                    h.get("nom_prenom", ""),
            "role":                    h.get("role") or "—",
            "societe":                 h.get("societe") or "—",
            "service":                 h.get("service") or "—",
            "domaine":                 h.get("domaine") or "—",
            "date_sensibilisation":    h.get("date_sensibilisation"),
            "sensibilisation_filename": h.get("sensibilisation_filename"),
            "status":                  status,
        }
        results.append(entry)
        if status != STATUS_CONFORME:
            ecarts.append({
                "systeme":              "ATT",
                "name":                 entry["name"],
                "role":                 entry["role"],
                "societe":              entry["societe"],
                "service":              entry["service"],
                "domaine":              entry["domaine"],
                "date_sensibilisation": entry["date_sensibilisation"],
                "detail":               f"{entry['role']} — {status}",
                "status":               status,
                "roles_lir":            [],
            })

    total     = len(results)
    conformes = sum(1 for r in results if r["status"] == STATUS_CONFORME)
    n_ecarts  = total - conformes

    ecarts = _group_ecarts(ecarts)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sensibilisations SMSI"
    ws.append(["Nom / Prénom", "Société", "Service", "Domaine", "Rôle", "Date sensibilisation", "Statut"])
    for c in ws[ws.max_row]:
        c.fill = HDR_F; c.font = HDR_T
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    fill_map = {STATUS_CONFORME: GREEN, STATUS_EXPIREE: ORANGE, STATUS_MANQUANTE: RED}
    for r in results:
        ws.append([r["name"], r["societe"], r["service"], r["domaine"],
                   r["role"], r["date_sensibilisation"] or "—", r["status"]])
        fill = fill_map.get(r["status"], RED)
        for c in ws[ws.max_row]:
            c.fill = fill

    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, 60)

    buf = io.BytesIO()
    wb.save(buf)

    return {
        "control_date": control_date,
        "lir_count":    len(habs),
        "resume": {
            "ATT": {"total": total, "conformes": conformes, "ecarts": n_ecarts},
            "total_ecarts": n_ecarts,
        },
        "results": results,
        "ecarts":  ecarts,
        "excel_bytes": buf.getvalue(),
    }


def compute_taux(result: dict) -> float:
    att    = result.get("resume", {}).get("ATT", {})
    total  = att.get("total", 0)
    ecarts = att.get("ecarts", 0)
    if total == 0:
        return 100.0
    return round((total - ecarts) / total * 100, 1)


def build_commentaire(result: dict) -> str:
    att    = result.get("resume", {}).get("ATT", {})
    nb     = att.get("ecarts", 0)
    total  = att.get("total", 0)
    date_s = result.get("control_date", "")
    if nb == 0:
        return f"Sensibilisations SMSI du {date_s} : {total} prestataire(s) contrôlé(s), aucun écart."
    return (f"Sensibilisations SMSI du {date_s} : {nb} écart(s) sur {total} prestataire(s) "
            f"(sensibilisation manquante ou expirée). Voir rapport Excel.")
